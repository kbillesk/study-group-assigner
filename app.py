"""
Study group optimisation: browser UI. Run with: python app.py
Then open http://127.0.0.1:5000/
"""

import os
import uuid
from io import BytesIO

from flask import Flask, render_template, request, send_file, url_for
from openpyxl import load_workbook

from study_groups import (
    build_csv,
    build_txt,
    get_next_study_group_sheet_name,
    get_prior_together_pairs,
    load_students_from_excel,
    solve,
)

app = Flask(__name__)
app.config["MAX_CONTENT_LENGTH"] = 16 * 1024 * 1024  # 16 MB max upload

# In-memory cache: token -> {"csv": str, "file": bytes, "groups": list of list of student dicts}
_download_cache: dict[str, dict] = {}


@app.route("/")
def index():
    return render_template("index.html")


@app.route("/run", methods=["POST"])
def run():
    if "file" not in request.files:
        return render_template("index.html", error="No file selected."), 400
    f = request.files["file"]
    if not f.filename or not f.filename.lower().endswith(".xlsx"):
        return render_template("index.html", error="Please upload an .xlsx file."), 400

    try:
        group_size = int(request.form.get("group_size", 0))
    except ValueError:
        group_size = 0
    if group_size <= 0:
        return render_template("index.html", error="Group size must be a positive integer."), 400

    same_sex = request.form.get("group_type") == "same_sex"

    try:
        data = f.read()
        students = load_students_from_excel(BytesIO(data))
    except Exception as e:
        return render_template("index.html", error=f"Could not read Excel file: {e}"), 400

    if len(students) == 0:
        return render_template("index.html", error="No valid students (B=sex, C=firstname, D=lastname) found in the file."), 400

    try:
        prior_pairs = get_prior_together_pairs(BytesIO(data))
    except Exception:
        prior_pairs = set()

    try:
        groups = solve(group_size, same_sex, students, prior_together_pairs=prior_pairs)
    except RuntimeError as e:
        return render_template("index.html", error=str(e)), 400

    groups_txt = build_txt(groups)
    csv_content = build_csv(groups)
    token = uuid.uuid4().hex
    original_filename = os.path.basename(f.filename or "study_groups_updated.xlsx")
    if not original_filename.lower().endswith(".xlsx"):
        original_filename = original_filename + ".xlsx" if "." not in original_filename else "study_groups_updated.xlsx"
    _download_cache[token] = {
        "csv": csv_content,
        "file": data,
        "groups": groups,
        "filename": original_filename,
    }

    return render_template(
        "result.html",
        token=token,
        num_groups=len(groups),
        num_students=len(students),
        group_type="Same-sex" if same_sex else "Mixed",
        groups_txt=groups_txt,
    )


@app.route("/download/<token>.csv")
def download_csv(token):
    if token not in _download_cache:
        return "Download expired or invalid.", 404
    csv_content = _download_cache[token]["csv"]
    return send_file(
        BytesIO(csv_content.encode("utf-8")),
        mimetype="text/csv; charset=utf-8",
        as_attachment=True,
        download_name="study_groups.csv",
    )


def _build_xlsx_with_groups_worksheet(file_bytes: bytes, groups: list) -> BytesIO:
    """Add a new worksheet study_group_1, study_group_2, ... Format: A=Group, B=Sex, C=Name."""
    wb = load_workbook(BytesIO(file_bytes), read_only=False)
    sheet_name = get_next_study_group_sheet_name(wb=wb)
    ws = wb.create_sheet(sheet_name)  # append at end so the first sheet stays the original
    ws.append(["Group", "Sex", "Name"])
    for i, group in enumerate(groups, start=1):
        for s in group:
            ws.append([i, s["sex"], s["name"]])
    out = BytesIO()
    wb.save(out)
    out.seek(0)
    return out


@app.route("/use_groups", methods=["POST"])
def use_groups():
    token = request.form.get("token")
    if not token or token not in _download_cache:
        return "Invalid or expired. Run the optimisation again.", 404
    entry = _download_cache[token]
    xlsx_io = _build_xlsx_with_groups_worksheet(entry["file"], entry["groups"])
    download_name = entry.get("filename", "study_groups_updated.xlsx")
    return send_file(
        xlsx_io,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=download_name,
    )


if __name__ == "__main__":
    app.run(host="127.0.0.1", port=5000, debug=True)

"""
Study group optimisation: load students from Excel, assign to groups via CP-SAT.
No web imports; use from Flask app or tests.
"""

import csv
import re
from datetime import datetime, timezone
from io import BytesIO, StringIO
from math import ceil
from typing import Any, BinaryIO, List, Set, Tuple, Union

from ortools.sat.python import cp_model

try:
    from openpyxl import load_workbook
except ImportError:
    load_workbook = None  # type: ignore

# Column indices in Excel (1-based for openpyxl): B=sex, C=firstname, D=lastname
COL_SEX = 2
COL_FIRSTNAME = 3
COL_LASTNAME = 4


def _normalize_sex(value: str) -> str:
    v = (value or "").strip().lower()
    if v in {"k", "f", "female", "kvinde"}:
        return "F"
    if v in {"m", "male", "mand"}:
        return "M"
    return (value or "").strip().upper()


def load_students_from_excel(
    filelike: Union[str, BinaryIO, BytesIO],
) -> List[dict]:
    """
    Load students from an Excel file. Columns: B=sex, C=firstname, D=lastname.
    Full name = firstname + " " + lastname (stripped). Only rows with sex F or M are included.
    """
    if load_workbook is None:
        raise ImportError("openpyxl is required; install with: pip install openpyxl")

    wb = load_workbook(read_only=True, data_only=True, filename=filelike)
    ws = wb.active
    students: List[dict] = []
    for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True), start=1):
        if not row or len(row) < max(COL_SEX, COL_FIRSTNAME, COL_LASTNAME):
            continue
        # Skip header: assume first row might be header (no integer in col B or names in C/D)
        sex_raw = row[COL_SEX - 1] if len(row) >= COL_SEX else None
        first = row[COL_FIRSTNAME - 1] if len(row) >= COL_FIRSTNAME else None
        last = row[COL_LASTNAME - 1] if len(row) >= COL_LASTNAME else None
        sex_raw = (sex_raw or "").strip() if sex_raw is not None else ""
        first = (first or "").strip() if first is not None else ""
        last = (last or "").strip() if last is not None else ""
        if not sex_raw and not first and not last:
            continue
        sex = _normalize_sex(sex_raw)
        if sex not in ("F", "M"):
            continue
        parts = [p for p in [first, last] if p]
        full_name = " ".join(parts) if parts else f"Student {row_idx}"
        students.append({
            "id": len(students),
            "name": full_name,
            "sex": sex,
        })
    wb.close()
    return students


# Worksheet naming: study_group_1, study_group_2, ...
STUDY_GROUP_SHEET_PREFIX = "study_group_"


def _study_group_sheet_numbers(wb: Any) -> List[int]:
    """Return sorted list of N for which sheet 'study_group_N' exists."""
    numbers = []
    pattern = re.compile(r"^study_group_(\d+)$", re.IGNORECASE)
    for name in wb.sheetnames:
        m = pattern.match(name.strip())
        if m:
            numbers.append(int(m.group(1)))
    return sorted(set(numbers))


def get_next_study_group_sheet_name(
    filelike: Union[str, BinaryIO, BytesIO, Any] = None,
    wb: Any = None,
) -> str:
    """
    Return the next worksheet name: study_group_1, study_group_2, etc.
    Pass either filelike (path or file) or an open workbook wb.
    """
    if wb is not None:
        numbers = _study_group_sheet_numbers(wb)
    else:
        if load_workbook is None:
            raise ImportError("openpyxl is required")
        wb = load_workbook(read_only=True, filename=filelike)
        numbers = _study_group_sheet_numbers(wb)
        wb.close()
    next_n = (max(numbers) + 1) if numbers else 1
    return f"{STUDY_GROUP_SHEET_PREFIX}{next_n}"


def get_prior_together_pairs(
    filelike: Union[str, BinaryIO, BytesIO],
) -> Set[Tuple[str, str]]:
    """
    Read all study_group_1, study_group_2, ... sheets and return pairs of names
    that have been in the same group before. Each pair is (name1, name2) with name1 < name2.
    """
    if load_workbook is None:
        raise ImportError("openpyxl is required")
    wb = load_workbook(read_only=True, data_only=True, filename=filelike)
    pairs: Set[Tuple[str, str]] = set()
    pattern = re.compile(r"^study_group_(\d+)$", re.IGNORECASE)
    for sheet_name in wb.sheetnames:
        if not pattern.match(sheet_name.strip()):
            continue
        ws = wb[sheet_name]
        # Columns: Group, Sex, Name (row 1 = header)
        by_group: dict = {}
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
            if not row or len(row) < 3:
                continue
            group_num, name = row[0], (row[2] or "").strip()
            if not name:
                continue
            try:
                g = int(group_num)
            except (TypeError, ValueError):
                continue
            by_group.setdefault(g, []).append(name)
        for names in by_group.values():
            for i, a in enumerate(names):
                for b in names[i + 1 :]:
                    pair = tuple(sorted([a, b]))
                    pairs.add(pair)
    wb.close()
    return pairs


def load_students_from_study_groups_sheet(
    filelike: Union[str, BinaryIO, BytesIO],
) -> List[dict]:
    """
    Load students from the latest study_group_N worksheet (Group, Sex, Name).
    Returns list of dicts with "id", "name", "sex", "group" (1-based group number).
    Use this to read a spreadsheet that was saved via "Use groups".
    """
    if load_workbook is None:
        raise ImportError("openpyxl is required; install with: pip install openpyxl")
    wb = load_workbook(read_only=True, data_only=True, filename=filelike)
    numbers = _study_group_sheet_numbers(wb)
    if not numbers:
        wb.close()
        return []
    sheet_name = f"{STUDY_GROUP_SHEET_PREFIX}{numbers[-1]}"
    ws = wb[sheet_name]
    students: List[dict] = []
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True), start=2):
        if not row or len(row) < 3:
            continue
        group_num, sex_raw, name = row[0], row[1], row[2]
        try:
            group_num = int(group_num) if group_num is not None else 0
        except (TypeError, ValueError):
            continue
        sex = _normalize_sex(str(sex_raw or ""))
        name = (name or "").strip() or f"Student {row_idx}"
        if sex not in ("F", "M"):
            continue
        students.append({
            "id": len(students),
            "name": name,
            "sex": sex,
            "group": group_num,
        })
    wb.close()
    return students


def solve(
    group_size: int,
    same_sex: bool,
    students: List[dict],
    prior_together_pairs: Union[Set[Tuple[str, str]], None] = None,
    time_limit_seconds: float = 30.0,
) -> List[List[dict]]:
    """
    Assign students to groups with at most group_size members per group.
    Number of groups G = ceil(n / group_size).
    same_sex=True => each group all F or all M.
    same_sex=False (mixed) => each group must have both sexes (no single-sex groups).
    prior_together_pairs: set of (name1, name2) that should be avoided in the same group (soft penalty).
    Returns list of groups; each group is a list of student dicts (same refs as input).
    """
    n = len(students)
    if n == 0:
        return [[]]
    if group_size <= 0:
        raise ValueError("group_size must be positive")

    prior_together_pairs = prior_together_pairs or set()
    G = ceil(n / group_size)
    model = cp_model.CpModel()

    # assign[s, g] = 1 if student s in group g
    assign = {}
    for s in students:
        for g in range(G):
            assign[(s["id"], g)] = model.NewBoolVar(f"s{s['id']}_g{g}")

    # Hard: each student in exactly one group
    for s in students:
        model.Add(sum(assign[(s["id"], g)] for g in range(G)) == 1)

    # Hard: each group size <= group_size
    for g in range(G):
        model.Add(
            sum(assign[(s["id"], g)] for s in students) <= group_size
        )

    females = [s for s in students if s["sex"] == "F"]
    males = [s for s in students if s["sex"] == "M"]

    if same_sex:
        # Same-sex: for each group, either all F or all M
        for g in range(G):
            female_count = sum(assign[(s["id"], g)] for s in females)
            male_count = sum(assign[(s["id"], g)] for s in males)
            is_female_only = model.NewBoolVar(f"female_only_g{g}")
            model.Add(male_count == 0).OnlyEnforceIf(is_female_only)
            model.Add(female_count == 0).OnlyEnforceIf(is_female_only.Not())
    else:
        # Mixed: avoid single-sex groups — each group must have at least 1 F and 1 M when we have enough of each
        if females and males and len(females) >= G and len(males) >= G:
            for g in range(G):
                female_count = sum(assign[(s["id"], g)] for s in females)
                male_count = sum(assign[(s["id"], g)] for s in males)
                model.Add(female_count >= 1)
                model.Add(male_count >= 1)

    # Soft: group sizes close to group_size
    penalties: List[Any] = []
    size_weight = 1
    for g in range(G):
        size_g = model.NewIntVar(0, n, f"size_g{g}")
        model.Add(size_g == sum(assign[(s["id"], g)] for s in students))
        dev = model.NewIntVar(0, group_size, f"dev_g{g}")
        model.AddAbsEquality(dev, size_g - group_size)
        penalties.append(dev * size_weight)

    # Soft (mixed only): 50-50 balance per group
    if not same_sex:
        balance_weight = 2
        for g in range(G):
            female_count = sum(assign[(s["id"], g)] for s in females)
            male_count = sum(assign[(s["id"], g)] for s in males)
            imbalance = model.NewIntVar(0, group_size, f"imbalance_g{g}")
            model.AddAbsEquality(imbalance, female_count - male_count)
            penalties.append(imbalance * balance_weight)

    # Soft: avoid putting prior-together pairs in the same group
    name_to_id = {s["name"].strip(): s["id"] for s in students}
    pair_weight = 10
    for (name1, name2) in prior_together_pairs:
        id1 = name_to_id.get(name1)
        id2 = name_to_id.get(name2)
        if id1 is None or id2 is None or id1 == id2:
            continue
        for g in range(G):
            both_in_g = model.NewBoolVar(f"prior_{id1}_{id2}_g{g}")
            model.AddMultiplicationEquality(both_in_g, [
                assign[(id1, g)],
                assign[(id2, g)],
            ])
            penalties.append(both_in_g * pair_weight)

    model.Minimize(sum(penalties))

    solver = cp_model.CpSolver()
    solver.parameters.max_time_in_seconds = time_limit_seconds
    status = solver.Solve(model)

    if status not in (cp_model.OPTIMAL, cp_model.FEASIBLE):
        raise RuntimeError(
            f"Solver could not find a solution: {solver.StatusName(status)}"
        )

    # Build groups: group[g] = list of students in group g
    groups: List[List[dict]] = [[] for _ in range(G)]
    for s in students:
        for g in range(G):
            if solver.Value(assign[(s["id"], g)]) == 1:
                groups[g].append(s)
                break
    return groups


def build_txt(groups: List[List[dict]]) -> str:
    """Build .txt report: one section per group with header and member lines."""
    lines = []
    for i, group in enumerate(groups, start=1):
        lines.append(f"=== Group {i} ===")
        lines.append(f"{'Name':<30} | {'Sex':<5}")
        lines.append("-" * 40)
        for s in group:
            lines.append(f"{s['name']:<30} | {s['sex']:<5}")
        lines.append(f"Total: {len(group)} students")
        lines.append("")
    return "\n".join(lines).strip() + "\n"


def build_csv(
    groups: List[List[dict]],
    timestamp: Union[str, datetime, None] = None,
    members_sep: str = ";",
) -> str:
    """
    Build .csv content: header timestamp,members; one row per group.
    timestamp defaults to now (ISO format) if not provided.
    """
    if timestamp is None:
        timestamp = datetime.now(timezone.utc).isoformat()
    elif isinstance(timestamp, datetime):
        timestamp = timestamp.isoformat()
    out = StringIO()
    writer = csv.writer(out)
    writer.writerow(["timestamp", "members"])
    for group in groups:
        members = members_sep.join(s["name"] for s in group)
        writer.writerow([timestamp, members])
    return out.getvalue()
